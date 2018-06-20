# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook(r'C:\Users\michael.dixon\Desktop\Template documents\doc generation spreadsheet - '
                   r'Rev 9  generate word document (Recovered).xlsm')

# Get sheet names
print(wb.sheetnames)

# Get a sheet by name
sheet = wb['Proposal']

# Print the sheet title
print(sheet.title)

# # Get currently active sheet
# # # anotherSheet = wb.active
# # #
# # # # Check `anotherSheet`
# # # anotherSheet

# Retrieve the value of a certain cell
print(sheet['A1'].value)

# Select element 'B2' of your sheet
c = sheet['B2']

# Retrieve the row number of your element
print(c.row)

# Retrieve the column letter of your element
print(c.column)

# Retrieve the coordinates of the cell
print(c.coordinate)

# Retrieve cell value
sheet.cell(row=1, column=2).value

# Print out values in column 2
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)


# Print row per row
for cellObj in sheet['A18':'H56']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)
      print('--- END ---')